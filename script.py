import json
import re
import unicodedata

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

JSON_FILE = "datos_extraidos_ia.json"
TEMPLATE_FILE = "Grupo Ovando.xlsx"
SHEET_NAME = "1. Datos"

# Filas de encabezado de anos (se actualizan solo si la celda no contiene formula).
HEADER_ROWS = [5, 42, 102]

# Regla de mapeo fijo de columnas.
YEAR_TO_COLUMN = {
    2019: "D",
    2020: "E",
    2021: "F",
    2022: "G",
    2023: "H",
    2024: "I",
}
COL_2025_ANUAL = "J"

FORCE_OVERWRITE_INJECTED_FORMULAS = True

# Filas que reciben dato duro (se usan tambien para limpiar columnas no mapeadas).
INJECTION_ROWS = [
    6, 8, 9, 13, 14, 15, 16, 17, 18, 20, 21, 24, 25, 26, 27,
    45, 46, 47, 48, 49, 50, 51,
    65, 66, 67, 68, 69, 70, 84,
    105, 106, 107, 108, 109,
    110, 111, 112, 113, 114, 115, 116, 117, 118,
    120, 121, 122,
    126, 127, 128,
]


def nombre_a_archivo(nombre):
    nombre = re.sub(r'[\\/:*?"<>|]', "", str(nombre or "")).strip()
    return (nombre or "Sin_Nombre") + ".xlsx"


def to_float(value):
    """Convierte a float de forma tolerante. Si falla, devuelve 0.0."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("$", "")
        if not cleaned:
            return 0.0
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


def get_alias_value(data, *keys, default=0.0):
    """
    Devuelve el valor del primer alias presente y no nulo en `data`.
    Si no encuentra ninguno, retorna `default`.
    """
    for key in keys:
        if key in data and data.get(key) is not None:
            return data.get(key)
    return default


def has_formula(value):
    return isinstance(value, str) and value.startswith("=")


def write_cell(ws, row, col, value, allow_formula_overwrite=False):
    """
    Escribe en celda respetando formulas (salvo excepciones explicitas).
    Retorna True si escribio, False si se omitio por proteccion de formula.
    """
    cell = ws.cell(row=row, column=col)
    if (
        has_formula(cell.value)
        and not allow_formula_overwrite
        and not FORCE_OVERWRITE_INJECTED_FORMULAS
    ):
        return False
    cell.value = value
    return True


def set_formula_cell(ws, formula, row=None, col=None, coord=None):
    """
    Escribe una formula de Excel forzando tipo formula (data_type='f').
    Acepta row/col o coord.
    """
    if coord is not None:
        cell = ws[coord]
    else:
        cell = ws.cell(row=row, column=col)

    if not isinstance(formula, str):
        formula = str(formula)
    formula = formula.lstrip()
    if not formula.startswith("="):
        formula = "=" + formula

    cell.value = formula
    if isinstance(cell.value, str) and cell.value.startswith("="):
        cell.data_type = "f"
    return cell


def _set_formula_if_empty(ws, row, col, formula):
    """Escribe formula solo cuando la celda esta vacia y no tiene formula."""
    cell = ws.cell(row=row, column=col)
    if has_formula(cell.value):
        return
    if cell.value is None or cell.value == "":
        cell.value = formula


def ensure_required_formulas(ws):
    """
    Completa formulas faltantes de filas calculadas en D:K.
    Solo rellena celdas vacias para no romper formulas/valores existentes.
    """
    col_start = column_index_from_string("D")
    col_end = column_index_from_string("K")

    for col in range(col_start, col_end + 1):
        col_l = get_column_letter(col)
        prev_l = get_column_letter(col - 1) if col > col_start else None

        formulas = {
            10: f"={col_l}8-{col_l}9",
            12: f"=SUM({col_l}13:{col_l}18)",
            19: f"={col_l}10-{col_l}12",
            29: f"=SUM({col_l}25:{col_l}28)",
            30: f"={col_l}19-{col_l}20+{col_l}21+{col_l}24-{col_l}29",
            44: f"=SUM({col_l}45:{col_l}63)",
            64: f"=SUM({col_l}65:{col_l}82)",
            83: f"=SUM({col_l}84:{col_l}94)",
            95: f"={col_l}44+{col_l}64+{col_l}83",
            97: f"={col_l}68",
            104: f"=SUM({col_l}105:{col_l}118)",
            119: f"=SUM({col_l}120:{col_l}122)",
            123: f"={col_l}104+{col_l}119",
            125: f"=SUM({col_l}126:{col_l}134)",
            135: f"={col_l}125",
            137: f"={col_l}123+{col_l}135",
        }

        if prev_l:
            formulas[98] = f"={col_l}97-{prev_l}97"
        else:
            formulas[98] = f"={col_l}97"

        for row, formula in formulas.items():
            _set_formula_if_empty(ws, row, col, formula)


def inject_native_formulas(ws, col):
    """
    Inyecta (overwrite) las formulas nativas del modelo para la columna actual.
    """
    col_l = get_column_letter(col)
    d_col = column_index_from_string("D")
    prev_l = get_column_letter(col - 1) if col > d_col else None

    formulas = {
        10: f"={col_l}8-{col_l}9",
        12: f"=SUM({col_l}13:{col_l}18)",
        19: f"={col_l}10-{col_l}12",
        29: f"=SUM({col_l}25:{col_l}28)",
        30: f"={col_l}19-{col_l}20+{col_l}21+{col_l}24-{col_l}29",
        44: f"=SUM({col_l}45:{col_l}63)",
        64: f"=SUM({col_l}65:{col_l}82)",
        83: f"=SUM({col_l}84:{col_l}94)",
        95: f"={col_l}44+{col_l}64+{col_l}83",
        97: f"={col_l}68",
        104: f"=SUM({col_l}105:{col_l}118)",
        119: f"=SUM({col_l}120:{col_l}122)",
        123: f"={col_l}104+{col_l}119",
        125: f"=SUM({col_l}126:{col_l}134)",
        135: f"={col_l}125",
        137: f"={col_l}123+{col_l}135",
    }

    if prev_l:
        formulas[98] = f"={col_l}97-{prev_l}97"
    else:
        formulas[98] = f"={col_l}97"

    for row, formula in formulas.items():
        write_cell(ws, row, col, formula, allow_formula_overwrite=True)


def clear_unmapped_columns(ws, mapped_cols):
    """
    Limpia columnas historicas no incluidas en el JSON actual.
    Se limpia solo D:J; la K se preserva para formulas automaticas del modelo.
    """
    d_col = column_index_from_string("D")
    j_col = column_index_from_string("J")

    for col in range(d_col, j_col + 1):
        if col in mapped_cols:
            continue

        for row in HEADER_ROWS:
            write_cell(ws, row, col, None, allow_formula_overwrite=True)

        for row in INJECTION_ROWS:
            write_cell(ws, row, col, 0.0, allow_formula_overwrite=True)


def resolve_target_column(periodo):
    """Resuelve la columna destino segun anio y tipo_periodo."""
    anio = int(to_float(periodo.get("anio", 0)))

    if anio in YEAR_TO_COLUMN:
        return column_index_from_string(YEAR_TO_COLUMN[anio]), anio

    if anio == 2025:
        # Por instruccion de negocio, 2025 siempre se inyecta en J.
        return column_index_from_string(COL_2025_ANUAL), anio

    return None, None


def inject_headers(ws, col, label):
    for row in HEADER_ROWS:
        write_cell(ws, row, col, label, allow_formula_overwrite=False)


def inject_estado_resultados(ws, col, periodo):
    er = periodo.get("estado_resultados") or {}

    ingresos = to_float(er.get("ingresos_operativos_netos"))
    costo_ventas = abs(to_float(er.get("costo_de_ventas")))
    gastos_operativos = abs(
        to_float(
            get_alias_value(
                er,
                "gastos_operativos",
                "gastos_operativos_totales",
            )
        )
    )
    gastos_generales = abs(to_float(er.get("gastos_generales")))
    gastos_arrendamientos = abs(to_float(er.get("gastos_por_arrendamientos")))
    servicios_externos_honorarios = abs(
        to_float(er.get("servicios_externos_y_honorarios"))
    )
    gastos_admin = abs(
        to_float(
            get_alias_value(
                er,
                "gastos_de_administracion",
                "gastos_de_administration",
            )
        )
    )
    gastos_venta = abs(to_float(er.get("gastos_de_venta")))
    gastos_personal = abs(to_float(er.get("gastos_de_personal")))
    otros_ingresos_operativos = abs(to_float(er.get("otros_ingresos_operativos")))
    otros_gastos_operativos = abs(to_float(er.get("otros_gastos_operativos")))
    otros_gastos_no_operativos = abs(to_float(er.get("otros_gastos_no_operativos")))
    otros_ingresos_no_operativos = abs(
        to_float(er.get("otros_ingresos_no_operativos"))
    )
    rif = to_float(er.get("resultado_financiero_neto"))
    isr_diferido = abs(to_float(er.get("isr_diferido")))
    isr_corriente = abs(to_float(er.get("isr_corriente")))
    provision_ptu = abs(to_float(er.get("provision_ptu")))
    total_impuestos_generico = abs(to_float(er.get("total_impuestos_generico")))
    # Anti-doble conteo con residuo:
    # fila 13 toma lo faltante para llegar al total de gastos_operativos.
    gastos_desgloses_sum = (
        gastos_admin
        + gastos_venta
        + gastos_personal
        + gastos_generales
        + gastos_arrendamientos
        + servicios_externos_honorarios
    )
    gastos_operativos_f13 = gastos_operativos - gastos_desgloses_sum
    if gastos_operativos_f13 < 0:
        gastos_operativos_f13 = 0.0
    gastos_generales_f14 = (
        gastos_generales
        + gastos_arrendamientos
        + servicios_externos_honorarios
    )
    otros_gastos_ingresos_op_f18 = (
        otros_gastos_operativos - otros_ingresos_operativos
    )

    # Fila 6 y 8: ingresos_operativos_netos.
    write_cell(ws, 6, col, ingresos)
    write_cell(ws, 8, col, ingresos)

    # Fila 9: costo de ventas en valor absoluto.
    write_cell(ws, 9, col, costo_ventas)
    write_cell(
        ws,
        10,
        col,
        f"={get_column_letter(col)}8-{get_column_letter(col)}9",
        allow_formula_overwrite=True,
    )
    write_cell(
        ws,
        12,
        col,
        f"=SUM({get_column_letter(col)}13:{get_column_letter(col)}18)",
        allow_formula_overwrite=True,
    )

    # Filas 13 a 18.
    write_cell(ws, 13, col, gastos_operativos_f13)
    write_cell(ws, 14, col, gastos_generales_f14)
    write_cell(ws, 15, col, gastos_admin)

    # Filas 16 a 18: desglose operativo.
    ws.cell(row=16, column=3).value = "Gastos de venta"
    ws.cell(row=17, column=3).value = "Gastos de personal"
    ws.cell(row=18, column=3).value = "Otros Gastos/Ingresos Op"
    write_cell(ws, 16, col, gastos_venta)
    write_cell(ws, 17, col, gastos_personal)
    write_cell(ws, 18, col, otros_gastos_ingresos_op_f18)

    # Filas 20 y 21: no operativos.
    ws.cell(row=20, column=3).value = "Otros Gastos No Op"
    ws.cell(row=21, column=3).value = "Otros Ingresos No Op"
    write_cell(ws, 20, col, otros_gastos_no_operativos)
    write_cell(ws, 21, col, otros_ingresos_no_operativos)
    write_cell(
        ws,
        19,
        col,
        f"={get_column_letter(col)}10-{get_column_letter(col)}12",
        allow_formula_overwrite=True,
    )

    # Fila 24: RIF inyectado como dato duro (sobrescribe formula por instruccion).
    write_cell(
        ws,
        24,
        col,
        rif,
        allow_formula_overwrite=True,
    )

    # Filas 25, 26 y 27: impuestos en absoluto.
    write_cell(ws, 25, col, isr_diferido)
    write_cell(ws, 26, col, isr_corriente)
    write_cell(ws, 27, col, provision_ptu)
    write_cell(
        ws,
        29,
        col,
        f"=SUM({get_column_letter(col)}25:{get_column_letter(col)}28)",
        allow_formula_overwrite=True,
    )

    # Fila 29: si 25,26,27 son 0 y llega total agrupado, inyectar valor duro.
    if (
        abs(isr_diferido) < 1e-9
        and abs(isr_corriente) < 1e-9
        and abs(provision_ptu) < 1e-9
        and abs(total_impuestos_generico) >= 1e-9
    ):
        write_cell(
            ws,
            29,
            col,
            total_impuestos_generico,
            allow_formula_overwrite=True,
        )

    # Fila 30: utilidad neta calculada con no operativos + RIF - impuestos.
    write_cell(
        ws,
        30,
        col,
        f"={get_column_letter(col)}19-{get_column_letter(col)}20+"
        f"{get_column_letter(col)}21+{get_column_letter(col)}24-"
        f"{get_column_letter(col)}29",
        allow_formula_overwrite=True,
    )


def inject_balance_general(ws, col, periodo):
    bg = periodo.get("balance_general") or {}
    activos = bg.get("activos") or {}
    pasivos = bg.get("pasivos") or {}
    capital = bg.get("capital_contable") or {}

    activos_circulante = activos.get("circulante") or {}
    activos_no_circulante = activos.get("no_circulante") or {}
    pasivo_cp = pasivos.get("corto_plazo") or {}
    pasivo_lp = pasivos.get("largo_plazo") or {}

    # Activo circulante (NO TOCAR fila 44).
    write_cell(
        ws,
        45,
        col,
        to_float(activos_circulante.get("efectivo_y_equivalentes")),
    )
    write_cell(
        ws,
        46,
        col,
        to_float(activos_circulante.get("cuentas_por_cobrar_clientes")),
    )
    write_cell(
        ws,
        47,
        col,
        to_float(activos_circulante.get("impuestos_a_favor_cp")),
    )
    write_cell(
        ws,
        48,
        col,
        to_float(activos_circulante.get("otros_activos_circulantes")),
    )
    write_cell(
        ws,
        49,
        col,
        to_float(activos_circulante.get("deudores_diversos_cp")),
    )
    write_cell(
        ws,
        50,
        col,
        to_float(activos_circulante.get("pagos_anticipados")),
    )
    ws.cell(row=51, column=3).value = "Inventarios"
    write_cell(
        ws,
        51,
        col,
        to_float(activos_circulante.get("inventarios")),
    )

    # Activo no circulante.
    ppe_transporte = to_float(activos_no_circulante.get("equipo_de_transporte"))
    ppe_computo = to_float(activos_no_circulante.get("equipo_de_computo"))
    ppe_mobiliario = to_float(
        activos_no_circulante.get("mobiliario_y_equipo_de_oficina")
    )
    ppe_depreciacion_abs = abs(
        to_float(
            get_alias_value(
                activos_no_circulante,
                "depreciacion_acumulada_historica",
                "depreciacion_acumulada",
            )
        )
    )
    ppe_neto = to_float(activos_no_circulante.get("propiedad_planta_y_equipo_neto"))

    write_cell(
        ws,
        65,
        col,
        ppe_transporte,
    )
    write_cell(
        ws,
        66,
        col,
        ppe_computo,
    )
    write_cell(
        ws,
        67,
        col,
        ppe_mobiliario,
    )
    # Depreciacion acumulada historica SIEMPRE se inyecta en negativo.
    write_cell(
        ws,
        68,
        col,
        -ppe_depreciacion_abs,
    )
    # Regla anti-doble conteo PPE:
    # si hay desglose en 65-68, no inyectar propiedad_planta_y_equipo_neto.
    # solo si NO hay desglose y PPE neto > 0, inyectar en fila 69.
    ppe_desglosado_sum = (
        abs(ppe_transporte) + abs(ppe_computo) + abs(ppe_mobiliario) + ppe_depreciacion_abs
    )
    if ppe_desglosado_sum > 1e-9:
        write_cell(ws, 69, col, 0.0)
    elif ppe_neto > 1e-9:
        ws.cell(row=69, column=3).value = "PPE Neto"
        write_cell(ws, 69, col, ppe_neto)
    else:
        write_cell(ws, 69, col, 0.0)

    ws.cell(row=70, column=3).value = "Activos Intangibles"
    write_cell(
        ws,
        70,
        col,
        to_float(activos_no_circulante.get("activos_intangibles_neto")),
    )

    # Activos diferidos (NO TOCAR fila 83).
    ws.cell(row=84, column=3).value = "Activos Diferidos"
    write_cell(
        ws,
        84,
        col,
        to_float(activos_no_circulante.get("activos_diferidos")),
    )

    # Pasivo circulante (NO TOCAR fila 104).
    pasivo_cp_map = [
        (105, "proveedores", "Proveedores"),
        (
            106,
            "impuestos_y_cuotas_por_pagar",
            "Impuestos y cuotas por pagar",
        ),
        (
            107,
            "otros_pasivos_corto_plazo",
            "Otros Pasivos CP",
        ),
        (108, "acreedores_diversos", "Acreedores diversos"),
        (109, "provisiones", "Provisiones"),
        (110, "anticipo_de_clientes", "Anticipo de Clientes"),
        (111, "deuda_financiera_cp", "Deuda financiera CP"),
    ]
    for row, key, label in pasivo_cp_map:
        ws.cell(row=row, column=3).value = label
        write_cell(ws, row, col, to_float(pasivo_cp.get(key)))

    # Limpia filas no utilizadas del bloque 112-118 para no arrastrar basura.
    for row in range(112, 119):
        write_cell(ws, row, col, 0.0)

    # Pasivo largo plazo (NO TOCAR filas 119 y 123).
    write_cell(ws, 120, col, to_float(pasivo_lp.get("dividendos_decretados")))
    write_cell(ws, 121, col, to_float(pasivo_lp.get("pasivo_por_arrendamiento")))
    write_cell(ws, 122, col, to_float(pasivo_lp.get("deuda_financiera_lp")))

    # Capital (NO TOCAR filas 125, 135 y 137).
    write_cell(ws, 126, col, to_float(capital.get("capital_social")))
    write_cell(
        ws,
        127,
        col,
        to_float(capital.get("utilidades_ejercicios_anteriores")),
    )
    write_cell(
        ws,
        128,
        col,
        to_float(capital.get("resultado_del_ejercicio_balance")),
    )


def repair_resumen_escenario(wb):
    """
    Repara desfasajes y referencias rotas en la hoja 4.Resumen-Escenario.
    """
    sheet_name = "4.Resumen-Escenario"
    datos_name = "1. Datos"
    if sheet_name not in wb.sheetnames:
        print(
            f"AVISO: no se encontro la hoja '{sheet_name}' para reparacion."
        )
        return
    if datos_name not in wb.sheetnames:
        print(
            f"AVISO: no se encontro la hoja '{datos_name}' para reparar '{sheet_name}'."
        )
        return

    ws = wb[sheet_name]
    ws_datos = wb[datos_name]

    row_depreciacion = _find_row_by_terms(
        ws_datos,
        include_terms=["depreciacion"],
        exclude_terms=["acumulada"],
        default_row=98,
        label_col=3,
    )
    if row_depreciacion == 98:
        # Respaldo por nombre historico en la plantilla.
        row_depreciacion = _find_row_by_terms(
            ws_datos,
            include_terms=["depreciacion", "periodo"],
            exclude_terms=["acumulada"],
            default_row=98,
            label_col=3,
        )

    # 1) Realineacion columna T (2024).
    set_formula_cell(ws, "='1. Datos'!I5", coord="T6")
    set_formula_cell(ws, "='1. Datos'!I19", coord="T11")
    set_formula_cell(ws, f"='1. Datos'!I{row_depreciacion}", coord="T18")

    # 2) Realineacion columna U (2025 base/proyeccion).
    set_formula_cell(ws, "='1. Datos'!J5", coord="U6")
    set_formula_cell(ws, "='1. Datos'!J19", coord="U11")
    set_formula_cell(ws, f"='1. Datos'!J{row_depreciacion}", coord="U18")

    # Ajuste explicito de referencias de 2.Cálculos (2) para 2024/2025.
    set_formula_cell(ws, "='2.Cálculos (2)'!H30", coord="T21")
    set_formula_cell(ws, "='2.Cálculos (2)'!I30", coord="U21")
    set_formula_cell(ws, "='2.Cálculos (2)'!H32", coord="T22")
    set_formula_cell(ws, "='2.Cálculos (2)'!I32", coord="U22")

    # 3) Curacion de errores: limpiar formulas rotas.
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if "#REF!" in cell.value or "[3]" in cell.value:
                cell.value = 0.0
                print(
                    "ALERTA: Se limpió referencia rota en la celda "
                    f"{cell.coordinate} de 4.Resumen-Escenario"
                )


def _normalize_text(value):
    if not isinstance(value, str):
        return ""
    text = value.strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )
    return text


def _find_row_by_labels(ws, labels, default_row, label_col=3):
    normalized_labels = {_normalize_text(lbl) for lbl in labels}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=label_col).value
        if _normalize_text(val) in normalized_labels:
            return row
    return default_row


def _find_rows_by_labels(ws, labels, label_col=2):
    normalized_labels = {_normalize_text(lbl) for lbl in labels}
    rows = []
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=label_col).value
        if _normalize_text(val) in normalized_labels:
            rows.append(row)
    return rows


def _find_row_by_terms(
    ws,
    include_terms,
    default_row,
    label_col=3,
    exclude_terms=None,
):
    include = [_normalize_text(t) for t in include_terms if t]
    exclude = [_normalize_text(t) for t in (exclude_terms or []) if t]

    for row in range(1, ws.max_row + 1):
        val = _normalize_text(ws.cell(row=row, column=label_col).value)
        if not val:
            continue
        if not all(term in val for term in include):
            continue
        if any(term in val for term in exclude):
            continue
        return row
    return default_row


def _find_row_contains_terms(ws, terms, default_row, label_col=1):
    normalized_terms = [_normalize_text(t) for t in terms if t]
    for row in range(1, ws.max_row + 1):
        val = _normalize_text(ws.cell(row=row, column=label_col).value)
        if not val:
            continue
        if any(term in val for term in normalized_terms):
            return row
    return default_row


def _find_value_col_for_label_row(ws, row, label_col=1):
    """
    Detecta la celda de valor a la derecha de la etiqueta.
    Prefiere la primera celda no vacia en las siguientes 3 columnas.
    """
    max_scan = min(ws.max_column, label_col + 3)
    for col in range(label_col + 1, max_scan + 1):
        val = ws.cell(row=row, column=col).value
        if val not in (None, ""):
            return col
    return label_col + 1


def repair_company_name_placeholders(wb, company_name):
    """
    Reemplaza textos hardcodeados de la empresa de plantilla por
    metadata.empresa_detectada en todas las hojas.
    """
    empresa = str(company_name or "").strip() or "Sin Nombre"
    replaced = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue

                text = value.strip()
                if not text or text.startswith("="):
                    continue

                normalized = _normalize_text(text)
                new_value = None

                # Caso ejecutivo: "Valoracion Negocio En ...".
                if "valoracion negocio en" in normalized:
                    new_value = f"Valoracion Negocio En {empresa}"
                # Casos de nombre hardcodeado de plantilla (Grupo Ovando / variantes).
                elif "ovando" in normalized:
                    if re.search(r"(?i)grupo\s+ovando", text):
                        # Si la celda es principalmente la razon social vieja, reemplaza todo.
                        if re.fullmatch(
                            r"(?is)\s*grupo\s+ovando[^\r\n]*",
                            text,
                        ):
                            new_value = empresa
                        else:
                            new_value = re.sub(
                                r"(?i)grupo\s+ovando",
                                empresa,
                                text,
                            )
                    else:
                        new_value = empresa

                if new_value is not None and new_value != value:
                    cell.value = new_value
                    replaced += 1
                    print(
                        "INFO: nombre empresa dinamico aplicado en "
                        f"{ws.title}!{cell.coordinate}"
                    )

    # Refuerzo explicito en anclas conocidas del modelo.
    if "1. Datos" in wb.sheetnames:
        wb["1. Datos"]["C2"].value = empresa
    if "RESUMEN" in wb.sheetnames:
        wb["RESUMEN"]["A1"].value = f"Valoracion Negocio En {empresa}"

    print(
        "INFO: reparacion de nombre de empresa completada. "
        f"Celdas actualizadas: {replaced}"
    )


def repair_dupont(wb):
    """
    Repara y reconecta la hoja Dupont con los totales de 1. Datos.
    """
    dupont_name = "Dupont"
    datos_name = "1. Datos"

    if dupont_name not in wb.sheetnames:
        print(f"AVISO: no se encontro la hoja '{dupont_name}' para reparacion.")
        return
    if datos_name not in wb.sheetnames:
        print(
            f"AVISO: no se encontro la hoja '{datos_name}' para reparar '{dupont_name}'."
        )
        return

    ws_dupont = wb[dupont_name]
    ws_datos = wb[datos_name]

    # Detecta filas reales de totales en 1. Datos (con fallback).
    row_activo_total = _find_row_by_labels(
        ws_datos,
        labels=["Activo Total", "Total Activos"],
        default_row=95,
    )
    row_pasivo_total = _find_row_by_labels(
        ws_datos,
        labels=["Pasivo Total", "Total Pasivos"],
        default_row=123,
    )
    row_capital_total = _find_row_by_labels(
        ws_datos,
        labels=[
            "Capital Total",
            "Total Capital Contable",
            "Capital Contable Total",
        ],
        default_row=135,
    )

    # Actualizacion dinamica ultimos 2 anos (G=2024, H=2025).
    set_formula_cell(ws_dupont, "='1. Datos'!I5", coord="G10")
    set_formula_cell(ws_dupont, "='1. Datos'!J5", coord="H10")

    set_formula_cell(ws_dupont, "='1. Datos'!I8", coord="G11")
    set_formula_cell(ws_dupont, "='1. Datos'!J8", coord="H11")

    set_formula_cell(ws_dupont, "='1. Datos'!I30", coord="G12")
    set_formula_cell(ws_dupont, "='1. Datos'!J30", coord="H12")

    set_formula_cell(ws_dupont, f"='1. Datos'!I{row_activo_total}", coord="G13")
    set_formula_cell(ws_dupont, f"='1. Datos'!J{row_activo_total}", coord="H13")

    set_formula_cell(ws_dupont, f"='1. Datos'!I{row_pasivo_total}", coord="G14")
    set_formula_cell(ws_dupont, f"='1. Datos'!J{row_pasivo_total}", coord="H14")

    set_formula_cell(ws_dupont, f"='1. Datos'!I{row_capital_total}", coord="G15")
    set_formula_cell(ws_dupont, f"='1. Datos'!J{row_capital_total}", coord="H15")

    # Curacion de referencias rotas en toda la hoja Dupont.
    for row in ws_dupont.iter_rows(
        min_row=1,
        max_row=ws_dupont.max_row,
        min_col=1,
        max_col=ws_dupont.max_column,
    ):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if "#REF!" in cell.value or "[3]" in cell.value:
                cell.value = 0.0
                print(
                    "ALERTA: Se limpió referencia rota en la celda "
                    f"{cell.coordinate} de Dupont"
                )


def repair_wacc(wb):
    """
    Repara conexion de deuda/capital y limpia referencias rotas en la hoja WACC.
    """
    wacc_name = "WACC"
    datos_name = "1. Datos"

    if wacc_name not in wb.sheetnames:
        print(f"AVISO: no se encontro la hoja '{wacc_name}' para reparacion.")
        return
    if datos_name not in wb.sheetnames:
        print(
            f"AVISO: no se encontro la hoja '{datos_name}' para reparar '{wacc_name}'."
        )
        return

    ws_wacc = wb[wacc_name]
    ws_datos = wb[datos_name]

    # Detecta filas reales en 1. Datos (con fallback).
    row_deuda_cp = _find_row_by_labels(
        ws_datos,
        labels=["Deuda financiera CP", "Deuda CP", "Deuda financiera corto plazo"],
        default_row=111,
    )
    row_deuda_lp = _find_row_by_labels(
        ws_datos,
        labels=["Deuda financiera LP", "Deuda LP", "Deuda financiera largo plazo"],
        default_row=122,
    )
    row_capital_total = _find_row_by_labels(
        ws_datos,
        labels=[
            "Capital Total",
            "Total Capital Contable",
            "Capital Contable Total",
        ],
        default_row=135,
    )

    deuda_formula = f"='1. Datos'!J{row_deuda_cp}+'1. Datos'!J{row_deuda_lp}"
    capital_formula = f"='1. Datos'!J{row_capital_total}"

    # Busca celdas de valores en pesos para Deuda y Capital (bloque inferior).
    debt_targets = []
    capital_targets = []
    debt_labels = {"dlp", "deuda", "deuda lp", "deuda financiera", "deuda total"}
    capital_labels = {"c", "capital", "capital contable", "equity"}

    for row in range(28, ws_wacc.max_row + 1):
        for col in range(1, ws_wacc.max_column):
            label = _normalize_text(ws_wacc.cell(row=row, column=col).value)
            if not label:
                continue
            next_coord = ws_wacc.cell(row=row, column=col + 1).coordinate
            if label in debt_labels:
                debt_targets.append(next_coord)
            if label in capital_labels:
                capital_targets.append(next_coord)

    # Si existe el bloque alterno (columna I) atado a Estructura de deuda, lo reconecta.
    if isinstance(ws_wacc["I30"].value, str) and "Estructura de deuda" in ws_wacc["I30"].value:
        debt_targets.append("I30")
    if isinstance(ws_wacc["I31"].value, str) and "Estructura de deuda" in ws_wacc["I31"].value:
        capital_targets.append("I31")

    # Fallback explicito si no se detectaron celdas de destino.
    if not debt_targets:
        debt_targets = ["C24"]
    if not capital_targets:
        capital_targets = ["C25"]

    # Deduplicar manteniendo orden.
    seen = set()
    debt_targets = [c for c in debt_targets if not (c in seen or seen.add(c))]
    seen = set()
    capital_targets = [c for c in capital_targets if not (c in seen or seen.add(c))]

    for coord in debt_targets:
        set_formula_cell(ws_wacc, deuda_formula, coord=coord)
    for coord in capital_targets:
        set_formula_cell(ws_wacc, capital_formula, coord=coord)

    # Curacion de referencias rotas.
    for row in ws_wacc.iter_rows(
        min_row=1,
        max_row=ws_wacc.max_row,
        min_col=1,
        max_col=ws_wacc.max_column,
    ):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if "#REF!" in cell.value or "[3]" in cell.value:
                cell.value = 0.0
                print(
                    "ALERTA: Se limpió referencia rota en la celda "
                    f"{cell.coordinate} de WACC"
                )


def repair_calculos_2(wb):
    """
    Reconecta entradas clave de 2.Cálculos (2) para 2024/2025 y limpia referencias rotas.
    """
    datos_name = "1. Datos"
    calc_name = None
    for sh in wb.sheetnames:
        sh_norm = _normalize_text(sh)
        if sh_norm.startswith("2.") and "calculos" in sh_norm:
            calc_name = sh
            break

    if calc_name is None:
        print("AVISO: no se encontro hoja de '2.Cálculos (2)' para reparacion.")
        return
    if datos_name not in wb.sheetnames:
        print(
            f"AVISO: no se encontro la hoja '{datos_name}' para reparar '{calc_name}'."
        )
        return

    ws_calc = wb[calc_name]
    ws_datos = wb[datos_name]

    # Detecta columnas 2024 y 2025 en la fila de anos.
    col_2024 = None
    col_2025 = None
    for row in range(1, min(25, ws_calc.max_row) + 1):
        found_2024 = None
        found_2025 = None
        for col in range(1, ws_calc.max_column + 1):
            val = ws_calc.cell(row=row, column=col).value
            val_norm = str(val).strip() if val is not None else ""
            if val == 2024 or val_norm == "2024":
                found_2024 = col
            if val == 2025 or val_norm == "2025":
                found_2025 = col
        if found_2024 and found_2025:
            col_2024, col_2025 = found_2024, found_2025
            break

    if col_2024 is None or col_2025 is None:
        # Fallback: H e I.
        col_2024, col_2025 = 8, 9

    # Filas fuente en 1. Datos.
    row_un = _find_row_by_labels(
        ws_datos,
        labels=[
            "UTILDIDAD O PERDIDA NETA DEL AÑO",
            "UTILDIDAD O PERDIDA NETA DEL ANO",
            "UTILIDAD NETA",
        ],
        default_row=30,
    )
    row_capital_total = _find_row_by_labels(
        ws_datos,
        labels=[
            "Capital Total",
            "Total Capital Contable",
            "Capital Contable Total",
        ],
        default_row=135,
    )
    row_ac = _find_row_by_terms(
        ws_datos,
        include_terms=["activo", "circulante"],
        default_row=44,
        label_col=3,
    )
    row_pc = _find_row_by_terms(
        ws_datos,
        include_terms=["pasivo", "circulante"],
        default_row=104,
        label_col=3,
    )
    row_ebit = _find_row_by_terms(
        ws_datos,
        include_terms=["utilidad", "operacion"],
        default_row=19,
        label_col=3,
    )
    row_dep = _find_row_by_terms(
        ws_datos,
        include_terms=["depreciacion"],
        exclude_terms=["acumulada"],
        default_row=98,
        label_col=3,
    )

    # Filas destino en 2.Cálculos (2), columna B.
    row_dest_un = _find_row_by_labels(
        ws_calc,
        labels=["Utilidad Neta"],
        default_row=6,
        label_col=2,
    )
    row_dest_capital = _find_row_by_labels(
        ws_calc,
        labels=["Capital Total"],
        default_row=7,
        label_col=2,
    )
    row_dest_ac = _find_row_by_labels(
        ws_calc,
        labels=["Activo Circulante"],
        default_row=17,
        label_col=2,
    )
    row_dest_pc = _find_row_by_labels(
        ws_calc,
        labels=["Pasivo Circulante"],
        default_row=18,
        label_col=2,
    )
    row_dest_ebit = _find_row_by_labels(
        ws_calc,
        labels=["Utilidad de Operación (EBIT)", "Utilidad de Operacion (EBIT)"],
        default_row=22,
        label_col=2,
    )
    rows_dest_dep = _find_rows_by_labels(
        ws_calc,
        labels=["Depreciación", "Depreciacion"],
        label_col=2,
    )
    if not rows_dest_dep:
        rows_dest_dep = [27, 31]

    # Inyeccion 2024/2025.
    base_map = [
        (row_dest_un, row_un),
        (row_dest_capital, row_capital_total),
        (row_dest_ac, row_ac),
        (row_dest_pc, row_pc),
        (row_dest_ebit, row_ebit),
    ]
    for dst_row, src_row in base_map:
        set_formula_cell(
            ws_calc,
            f"='1. Datos'!I{src_row}",
            row=dst_row,
            col=col_2024,
        )
        set_formula_cell(
            ws_calc,
            f"='1. Datos'!J{src_row}",
            row=dst_row,
            col=col_2025,
        )

    for dst_row in rows_dest_dep:
        set_formula_cell(
            ws_calc,
            f"='1. Datos'!I{row_dep}",
            row=dst_row,
            col=col_2024,
        )
        set_formula_cell(
            ws_calc,
            f"='1. Datos'!J{row_dep}",
            row=dst_row,
            col=col_2025,
        )

    # Curacion de referencias rotas.
    for row in ws_calc.iter_rows(
        min_row=1,
        max_row=ws_calc.max_row,
        min_col=1,
        max_col=ws_calc.max_column,
    ):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if "#REF!" in cell.value or "[3]" in cell.value:
                cell.value = 0.0
                print(
                    "ALERTA: Se limpió referencia rota en la celda "
                    f"{cell.coordinate} de {calc_name}"
                )


def repair_razones_financieras(wb):
    """
    Reconecta inputs base de KPIs en Razones financieras y limpia referencias rotas.
    """
    rf_name = "Razones financieras"
    datos_name = "1. Datos"

    if rf_name not in wb.sheetnames:
        print(f"AVISO: no se encontro la hoja '{rf_name}' para reparacion.")
        return
    if datos_name not in wb.sheetnames:
        print(
            f"AVISO: no se encontro la hoja '{datos_name}' para reparar '{rf_name}'."
        )
        return

    ws_rf = wb[rf_name]
    ws_datos = wb[datos_name]

    # Detecta filas reales de totales en 1. Datos.
    row_activo_total = _find_row_by_labels(
        ws_datos,
        labels=["Activo Total", "Total Activos"],
        default_row=95,
    )
    row_pasivo_total = _find_row_by_labels(
        ws_datos,
        labels=["Pasivo Total", "Total Pasivos"],
        default_row=123,
    )
    row_capital_total = _find_row_by_labels(
        ws_datos,
        labels=[
            "Capital Total",
            "Total Capital Contable",
            "Capital Contable Total",
        ],
        default_row=135,
    )

    # Detecta fila DATOS y ultimas 2 columnas de periodo en la serie historica.
    row_datos = _find_row_by_labels(
        ws_rf,
        labels=["DATOS"],
        default_row=4,
        label_col=2,
    )
    period_cols = []
    for col in range(3, ws_rf.max_column + 1):
        val = ws_rf.cell(row=row_datos, column=col).value
        if val is not None and val != "":
            period_cols.append(col)

    if len(period_cols) >= 2:
        col_2024, col_2025 = period_cols[-2], period_cols[-1]
    else:
        # Fallback pedido: si no se detecta dinamicamente, usar G y H.
        col_2024, col_2025 = 7, 8

    # Detecta filas clave en Razones financieras (columna B).
    row_ventas = _find_row_by_labels(
        ws_rf,
        labels=["Ventas Netas"],
        default_row=5,
        label_col=2,
    )
    row_un = _find_row_by_labels(
        ws_rf,
        labels=["UN (Utilidad Neta)", "Utilidad Neta", "UN"],
        default_row=6,
        label_col=2,
    )

    rows_activo = _find_rows_by_labels(
        ws_rf,
        labels=["Activo Total", "AT", "AT (Activo Total)"],
        label_col=2,
    ) or [7]
    rows_pasivo = _find_rows_by_labels(
        ws_rf,
        labels=["Pasivo Total", "PT", "PT (Pasivo Total)"],
        label_col=2,
    ) or [8]
    rows_capital = _find_rows_by_labels(
        ws_rf,
        labels=["Capital Contable", "CC", "CC (Capital Contable)"],
        label_col=2,
    ) or [9]

    # Inyeccion de ultimos 2 periodos.
    set_formula_cell(ws_rf, "='1. Datos'!I5", row=row_datos, col=col_2024)
    set_formula_cell(ws_rf, "='1. Datos'!J5", row=row_datos, col=col_2025)

    set_formula_cell(ws_rf, "='1. Datos'!I8", row=row_ventas, col=col_2024)
    set_formula_cell(ws_rf, "='1. Datos'!J8", row=row_ventas, col=col_2025)

    set_formula_cell(ws_rf, "='1. Datos'!I30", row=row_un, col=col_2024)
    set_formula_cell(ws_rf, "='1. Datos'!J30", row=row_un, col=col_2025)

    for row in rows_activo:
        set_formula_cell(
            ws_rf,
            f"='1. Datos'!I{row_activo_total}",
            row=row,
            col=col_2024,
        )
        set_formula_cell(
            ws_rf,
            f"='1. Datos'!J{row_activo_total}",
            row=row,
            col=col_2025,
        )

    for row in rows_pasivo:
        set_formula_cell(
            ws_rf,
            f"='1. Datos'!I{row_pasivo_total}",
            row=row,
            col=col_2024,
        )
        set_formula_cell(
            ws_rf,
            f"='1. Datos'!J{row_pasivo_total}",
            row=row,
            col=col_2025,
        )

    for row in rows_capital:
        set_formula_cell(
            ws_rf,
            f"='1. Datos'!I{row_capital_total}",
            row=row,
            col=col_2024,
        )
        set_formula_cell(
            ws_rf,
            f"='1. Datos'!J{row_capital_total}",
            row=row,
            col=col_2025,
        )

    # Subcuentas necesarias para evitar #DIV/0! en KPIs.
    row_ac = _find_row_by_terms(
        ws_datos,
        include_terms=["activo", "circulante"],
        default_row=44,
        label_col=3,
    )
    row_pc = _find_row_by_terms(
        ws_datos,
        include_terms=["pasivo", "circulante"],
        default_row=104,
        label_col=3,
    )
    row_cv = _find_row_by_terms(
        ws_datos,
        include_terms=["costo", "ventas"],
        default_row=9,
        label_col=3,
    )
    row_cxc = _find_row_by_terms(
        ws_datos,
        include_terms=["cuentas", "cobrar"],
        default_row=46,
        label_col=3,
    )
    row_cxp = _find_row_by_terms(
        ws_datos,
        include_terms=["proveedor"],
        default_row=105,
        label_col=3,
    )
    row_inv = _find_row_by_terms(
        ws_datos,
        include_terms=["inventario"],
        default_row=51,
        label_col=3,
    )
    row_ub = _find_row_by_terms(
        ws_datos,
        include_terms=["utilidad", "bruta"],
        default_row=10,
        label_col=3,
    )
    row_uo = _find_row_by_terms(
        ws_datos,
        include_terms=["utilidad", "operacion"],
        default_row=19,
        label_col=3,
    )

    row_dest_ac = _find_row_by_labels(
        ws_rf,
        labels=["AC (Activo Circulante)"],
        default_row=10,
        label_col=2,
    )
    row_dest_pc = _find_row_by_labels(
        ws_rf,
        labels=["PC (Pasivo Circulante)"],
        default_row=11,
        label_col=2,
    )
    row_dest_cv = _find_row_by_labels(
        ws_rf,
        labels=["CV (Costo de Ventas)"],
        default_row=16,
        label_col=2,
    )
    row_dest_cxc = _find_row_by_labels(
        ws_rf,
        labels=["CxC Cuentas por Cobrar"],
        default_row=17,
        label_col=2,
    )
    row_dest_cxp = _find_row_by_labels(
        ws_rf,
        labels=["CxP Cuentas por Pagar"],
        default_row=18,
        label_col=2,
    )
    row_dest_inv = _find_row_by_labels(
        ws_rf,
        labels=["Inventarios"],
        default_row=21,
        label_col=2,
    )
    row_dest_ub = _find_row_by_labels(
        ws_rf,
        labels=["UB (Utilidad Bruta)"],
        default_row=19,
        label_col=2,
    )
    row_dest_uo = _find_row_by_labels(
        ws_rf,
        labels=["UO (Utilidad Operativa)"],
        default_row=20,
        label_col=2,
    )

    subaccount_map = [
        (row_dest_ac, row_ac),
        (row_dest_pc, row_pc),
        (row_dest_cv, row_cv),
        (row_dest_cxc, row_cxc),
        (row_dest_cxp, row_cxp),
        (row_dest_inv, row_inv),
        (row_dest_ub, row_ub),
        (row_dest_uo, row_uo),
    ]
    for dst_row, src_row in subaccount_map:
        set_formula_cell(
            ws_rf,
            f"='1. Datos'!I{src_row}",
            row=dst_row,
            col=col_2024,
        )
        set_formula_cell(
            ws_rf,
            f"='1. Datos'!J{src_row}",
            row=dst_row,
            col=col_2025,
        )

    # Curacion de referencias rotas.
    for row in ws_rf.iter_rows(
        min_row=1,
        max_row=ws_rf.max_row,
        min_col=1,
        max_col=ws_rf.max_column,
    ):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if "#REF!" in cell.value or "[3]" in cell.value:
                cell.value = 0.0
                print(
                    "ALERTA: Se limpió referencia rota en la celda "
                    f"{cell.coordinate} de Razones financieras"
                )


def repair_resumen_final(wb):
    """
    Reconecta KPI ejecutivos en hoja RESUMEN y limpia referencias rotas.
    """
    resumen_name = "RESUMEN"
    datos_name = "1. Datos"

    if resumen_name not in wb.sheetnames:
        print(f"AVISO: no se encontro la hoja '{resumen_name}' para reparacion.")
        return
    if datos_name not in wb.sheetnames:
        print(
            f"AVISO: no se encontro la hoja '{datos_name}' para reparar '{resumen_name}'."
        )
        return

    ws_resumen = wb[resumen_name]
    ws_datos = wb[datos_name]

    # Filas detectadas en 1. Datos (con fallback).
    row_activo_total = _find_row_by_labels(
        ws_datos,
        labels=["Activo Total", "Total Activos"],
        default_row=95,
    )
    # Busqueda estricta de EBITDA en la seccion superior de 1. Datos.
    # Regla: etiqueta exacta "EBITDA" (sin texto adicional), maximo primeras 100 filas.
    max_row_to_scan = 100
    row_ebitda = 97  # fallback forzoso
    scan_limit = min(ws_datos.max_row, max_row_to_scan)
    for row in range(1, scan_limit + 1):
        label = _normalize_text(ws_datos.cell(row=row, column=3).value)
        if label == "ebitda":
            row_ebitda = row
            break

    # Buscar etiquetas en RESUMEN.
    row_ebitda_2025 = _find_row_by_labels(
        ws_resumen,
        labels=["EBITDA 2025"],
        default_row=6,
        label_col=1,
    )
    row_ebitda_ultimo = _find_row_by_labels(
        ws_resumen,
        labels=["EBITDA ULTIMO AÑO", "EBITDA ULTIMO ANO"],
        default_row=12,
        label_col=1,
    )
    row_ebitda_generic = _find_row_by_labels(
        ws_resumen,
        labels=["EBITDA"],
        default_row=row_ebitda_2025,
        label_col=1,
    )
    if row_ebitda_generic == row_ebitda_2025 and _normalize_text(
        ws_resumen.cell(row=row_ebitda_generic, column=1).value
    ) != "ebitda":
        row_ebitda_generic = _find_row_contains_terms(
            ws_resumen,
            terms=["ebitda"],
            default_row=row_ebitda_2025,
            label_col=1,
        )

    row_valor_contable = _find_row_by_labels(
        ws_resumen,
        labels=["Valor Neto Contable", "Capital Invertido"],
        default_row=5,
        label_col=1,
    )
    if row_valor_contable == 5 and _normalize_text(
        ws_resumen.cell(row=row_valor_contable, column=1).value
    ) not in {"valor neto contable", "capital invertido"}:
        row_valor_contable = _find_row_contains_terms(
            ws_resumen,
            terms=["valor neto contable", "capital invertido"],
            default_row=5,
            label_col=1,
        )

    # Detectar columna valor (normalmente B).
    ebitda_rows = sorted({row_ebitda_2025, row_ebitda_ultimo, row_ebitda_generic})
    for row in ebitda_rows:
        val_col = _find_value_col_for_label_row(ws_resumen, row, label_col=1)
        set_formula_cell(
            ws_resumen,
            f"='1. Datos'!J{row_ebitda}",
            row=row,
            col=val_col,
        )

    valor_col = _find_value_col_for_label_row(ws_resumen, row_valor_contable, label_col=1)
    set_formula_cell(
        ws_resumen,
        f"='1. Datos'!J{row_activo_total}",
        row=row_valor_contable,
        col=valor_col,
    )

    # Curacion de referencias rotas.
    for row in ws_resumen.iter_rows(
        min_row=1,
        max_row=ws_resumen.max_row,
        min_col=1,
        max_col=ws_resumen.max_column,
    ):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if "#REF!" in cell.value or "[3]" in cell.value:
                cell.value = 0.0
                print(
                    "ALERTA: Se limpió referencia rota en la celda "
                    f"{cell.coordinate} de RESUMEN"
                )


def inyectar_datos_financieros(json_path, template_path, output_path):
    with open(json_path, encoding="utf-8-sig") as f:
        data = json.load(f)

    periodos = data.get("datos_financieros") or []
    empresa = data.get("metadata", {}).get("empresa_detectada", "Sin Nombre")

    if not periodos:
        raise ValueError(
            "El JSON no contiene periodos en 'datos_financieros'."
        )

    print(f"Empresa:  {empresa}")
    print(f"Periodos recibidos: {[p.get('anio') for p in periodos]}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb[SHEET_NAME]
    ensure_required_formulas(ws)

    mapped_cols = set()
    for periodo in periodos:
        col, _ = resolve_target_column(periodo)
        if col is not None:
            mapped_cols.add(col)
    clear_unmapped_columns(ws, mapped_cols)

    used_cols = {}
    skipped = 0

    for periodo in periodos:
        col, label = resolve_target_column(periodo)
        if col is None:
            print(
                f"AVISO: se omite periodo anio={periodo.get('anio')} "
                f"tipo_periodo={periodo.get('tipo_periodo')} (sin columna definida)."
            )
            skipped += 1
            continue

        prev = used_cols.get(col)
        if prev is not None:
            print(
                f"AVISO: columna {col} ya usada por anio={prev}. "
                f"Se sobrescribe con anio={periodo.get('anio')}."
            )
        used_cols[col] = periodo.get("anio")

        inject_headers(ws, col, label)
        inject_native_formulas(ws, col)
        inject_estado_resultados(ws, col, periodo)
        inject_balance_general(ws, col, periodo)

    repair_calculos_2(wb)
    repair_resumen_escenario(wb)
    repair_dupont(wb)
    repair_wacc(wb)
    repair_razones_financieras(wb)
    repair_resumen_final(wb)
    repair_company_name_placeholders(wb, empresa)

    wb.save(output_path)
    try:
        check_wb = openpyxl.load_workbook(output_path, data_only=False, read_only=True)
        check_value = None
        if "RESUMEN" in check_wb.sheetnames:
            check_value = check_wb["RESUMEN"]["B5"].value
        print(f"CHECK POST-SAVE RESUMEN!B5 = {check_value!r}")
        check_wb.close()
    except Exception as exc:
        print(f"CHECK POST-SAVE ERROR: {exc}")

    print(
        f"Listo. Archivo generado en '{output_path}'. "
        f"Columnas cargadas: {sorted(used_cols.keys())}. "
        f"Periodos omitidos: {skipped}."
    )


def main():
    with open(JSON_FILE, encoding="utf-8") as f:
        empresa = (
            json.load(f).get("metadata", {}).get("empresa_detectada", "Sin Nombre")
        )
    inyectar_datos_financieros(
        JSON_FILE,
        TEMPLATE_FILE,
        nombre_a_archivo(empresa),
    )


if __name__ == "__main__":
    main()
