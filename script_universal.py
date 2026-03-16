"""
script_universal.py
===================
Motor de Inyeccion Dinamica Universal.

Cambios clave vs script.py:
  - LEFT-ALIGNED: anio 1 -> col D, anio N -> col D+N-1 (sin limite de columnas).
  - col_proj = col_last + 1  (columna proyectada, siempre dinamica).
  - col_helper = col_proj + 1 (anualizacion: col_last / mes_cierre).
  - Fila 4: formulas YoY inyectadas dinamicamente para cada columna.
  - Filas 5/42/102: headers con valor real por anio.
  - CAGR filas 192-213: N dinamico = col_last - col_first.
  - repair_*: backtracking celda-a-celda, soporte N anios ilimitado.
  - Soporta columnas AA, AB, ... via openpyxl get_column_letter.
"""

import json
import re
import unicodedata

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
JSON_FILE = "datos_extraidos_ia.json"
TEMPLATE_FILE = "Grupo Ovando.xlsx"
SHEET_NAME = "1. Datos"

# Filas de encabezado de anios en "1. Datos"
HEADER_ROWS = [5, 42, 102]

# Columna inicial de datos (siempre D = indice 4)
COL_FIRST_IDX = 4
COL_FIRST = "D"

# Filas de inyeccion directa: P&L y Balance
PL_INJECT_ROWS = [6, 8, 9, 13, 14, 15, 16, 17, 18, 20, 21, 24, 25, 26, 27]
BAL_INJECT_ROWS = [
    45, 46, 47, 48, 49, 50, 51,
    65, 66, 67, 68, 69, 70, 84,
    105, 106, 107, 108, 109, 110, 111, 112, 113, 114, 115, 116, 117, 118,
    120, 121, 122,
    126, 127, 128,
]
INJECTION_ROWS = PL_INJECT_ROWS + BAL_INJECT_ROWS

FORCE_OVERWRITE_INJECTED_FORMULAS = True

# Filas de subtotales/formulas nativas en "1. Datos"
# Clave: fila; valor: funcion(col_letter, prev_col_letter_or_None) -> formula_str
def _subtotal_formulas(col_l, prev_l):
    return {
        10:  f"={col_l}8-{col_l}9",
        12:  f"=SUM({col_l}13:{col_l}18)",
        19:  f"={col_l}10-{col_l}12",
        22:  f"={col_l}19+ABS({col_l}98)",
        29:  f"=SUM({col_l}25:{col_l}28)",
        30:  f"={col_l}19+{col_l}24-{col_l}29",
        44:  f"=SUM({col_l}45:{col_l}63)",
        64:  f"=SUM({col_l}65:{col_l}82)",
        83:  f"=SUM({col_l}84:{col_l}94)",
        95:  f"={col_l}44+{col_l}64+{col_l}83",
        97:  f"={col_l}68",
        98:  f"={col_l}97-{prev_l}97" if prev_l else f"={col_l}97",
        104: f"=SUM({col_l}105:{col_l}118)",
        119: f"=SUM({col_l}120:{col_l}122)",
        123: f"={col_l}104+{col_l}119",
        125: f"=SUM({col_l}126:{col_l}134)",
        135: f"={col_l}125",
        137: f"={col_l}123+{col_l}135",
    }


# ---------------------------------------------------------------------------
# Utilidades basicas
# ---------------------------------------------------------------------------

def to_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("$", "")
        if not cleaned:
            return 0.0
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


def get_alias_value(data, *keys, default=0.0):
    for key in keys:
        if key in data and data.get(key) is not None:
            return data.get(key)
    return default


def has_formula(value):
    return isinstance(value, str) and value.startswith("=")


def write_cell(ws, row, col, value, allow_formula_overwrite=False):
    cell = ws.cell(row=row, column=col)
    if (
        has_formula(cell.value)
        and not allow_formula_overwrite
        and not FORCE_OVERWRITE_INJECTED_FORMULAS
    ):
        return False
    cell.value = value
    return True


def set_formula_cell(ws, formula, row=None, col=None, coord=None):
    if coord is not None:
        cell = ws[coord]
    else:
        cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell  # no se puede escribir en celda fusionada secundaria
    if not isinstance(formula, str):
        formula = str(formula)
    formula = formula.lstrip()
    if not formula.startswith("="):
        formula = "=" + formula
    cell.value = formula
    if isinstance(cell.value, str) and cell.value.startswith("="):
        cell.data_type = "f"
    return cell


def nombre_a_archivo(nombre):
    nombre = re.sub(r'[\\/:*?"<>|]', "", str(nombre or "")).strip()
    return (nombre or "Sin_Nombre") + ".xlsx"


def _normalize_text(value):
    if not isinstance(value, str):
        return ""
    text = value.strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )
    return text


def _find_row_by_labels(ws, labels, default_row, label_col=3):
    normalized = {_normalize_text(lbl) for lbl in labels}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=label_col).value
        if _normalize_text(val) in normalized:
            return row
    return default_row


def _find_rows_by_labels(ws, labels, label_col=2):
    normalized = {_normalize_text(lbl) for lbl in labels}
    rows = []
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=label_col).value
        if _normalize_text(val) in normalized:
            rows.append(row)
    return rows


def _find_row_by_terms(ws, include_terms, default_row, label_col=3, exclude_terms=None):
    include = [_normalize_text(t) for t in include_terms if t]
    exclude = [_normalize_text(t) for t in (exclude_terms or []) if t]
    for row in range(1, ws.max_row + 1):
        val = _normalize_text(ws.cell(row=row, column=label_col).value)
        if not val:
            continue
        if not all(term in val for term in include):
            continue
        if any(term in val for term in exclude):
            continue
        return row
    return default_row


def _find_row_contains_terms(ws, terms, default_row, label_col=1):
    normalized = [_normalize_text(t) for t in terms if t]
    for row in range(1, ws.max_row + 1):
        val = _normalize_text(ws.cell(row=row, column=label_col).value)
        if not val:
            continue
        if any(term in val for term in normalized):
            return row
    return default_row


def _find_value_col_for_label_row(ws, row, label_col=1):
    max_scan = min(ws.max_column, label_col + 3)
    for col in range(label_col + 1, max_scan + 1):
        val = ws.cell(row=row, column=col).value
        if val not in (None, ""):
            return col
    return label_col + 1


def _find_calculos_sheet_name(wb):
    for sh in wb.sheetnames:
        sh_norm = _normalize_text(sh)
        if sh_norm.startswith("2.") and "calculos" in sh_norm:
            return sh
    return None


# ---------------------------------------------------------------------------
# Constructor de columnas dinamicas
# ---------------------------------------------------------------------------

def _build_cols(periodos):
    """
    Calcula los indices/letras de columnas para N periodos LEFT-ALIGNED.

    Retorna dict con:
      n           : cantidad de periodos
      first_idx   : indice col D = 4
      last_idx    : indice ultima col de datos reales
      proj_idx    : indice col proyectada (last + 1)
      helper_idx  : indice col helper anualizacion (last + 2)
      first/last/proj/helper: letras correspondientes
      data_cols   : lista de (idx, letra) para cols D..col_last
    """
    n = len(periodos)
    first_idx = COL_FIRST_IDX
    last_idx = first_idx + n - 1
    proj_idx = last_idx + 1
    helper_idx = proj_idx + 1
    return {
        "n": n,
        "first_idx": first_idx,
        "last_idx": last_idx,
        "proj_idx": proj_idx,
        "helper_idx": helper_idx,
        "first": get_column_letter(first_idx),
        "last": get_column_letter(last_idx),
        "proj": get_column_letter(proj_idx),
        "helper": get_column_letter(helper_idx),
        "data_cols": [
            (first_idx + i, get_column_letter(first_idx + i))
            for i in range(n)
        ],
    }


# ---------------------------------------------------------------------------
# Inyeccion de headers y formulas base en "1. Datos"
# ---------------------------------------------------------------------------

def inject_native_formulas(ws, col_idx):
    """Escribe/sobreescribe formulas de subtotales para la columna dada."""
    col_l = get_column_letter(col_idx)
    prev_l = get_column_letter(col_idx - 1) if col_idx > COL_FIRST_IDX else None
    for row, formula in _subtotal_formulas(col_l, prev_l).items():
        write_cell(ws, row, col_idx, formula, allow_formula_overwrite=True)


def inject_headers(ws, col_idx, label):
    """Escribe la etiqueta de anio en las filas de encabezado."""
    for row in HEADER_ROWS:
        write_cell(ws, row, col_idx, label, allow_formula_overwrite=False)


def inject_datos_year_headers(ws, cols, periodos):
    """
    Escribe los encabezados de anio en filas 5, 42 y 102.
    Fila 5: valor entero del anio (hardcoded por anio real).
    Filas 42 y 102: referencia a fila 5.
    """
    for i, (col_idx, col_l) in enumerate(cols["data_cols"]):
        periodo = periodos[i]
        anio = int(to_float(periodo.get("anio", 0)))
        tipo = str(periodo.get("tipo_periodo", "")).upper()
        # Etiqueta de fila 5
        if tipo == "PARCIAL":
            mes = int(periodo.get("mes_cierre") or 12)
            label5 = f"{anio} ({mes}m)"
        else:
            label5 = anio
        ws.cell(row=5, column=col_idx).value = label5
        # Filas 42 y 102 referencian fila 5
        ws.cell(row=42, column=col_idx).value = f"={col_l}5"
        ws.cell(row=102, column=col_idx).value = f"={col_l}5"

    # Columna proyectada
    ultimo = periodos[-1]
    anio_u = int(to_float(ultimo.get("anio", 0)))
    tipo_u = str(ultimo.get("tipo_periodo", "")).upper()
    proj_idx = cols["proj_idx"]
    proj_l = cols["proj"]
    label_proj = f"{anio_u} Proyectado" if tipo_u == "PARCIAL" else f"{anio_u} Anual"
    ws.cell(row=5, column=proj_idx).value = label_proj
    ws.cell(row=42, column=proj_idx).value = f"={proj_l}5"
    ws.cell(row=102, column=proj_idx).value = f"={proj_l}5"


def inject_datos_yoy_row4(ws, cols, periodos):
    """
    Fila 4: formulas de crecimiento YoY.
    Patron: col_i4 = col_i6 / col_prev6 - 1
    Se omite para columnas PARCIAL (solo el col_proj siempre recibe formula).
    La celda tras col_proj recibe =AVERAGE(segunda_col4:col_proj4).
    """
    # Determinar segunda col (primera que recibe formula YoY)
    data_cols = cols["data_cols"]
    if len(data_cols) < 2:
        # Solo 1 anio: no hay YoY historico
        pass
    else:
        for i in range(1, len(data_cols)):
            col_idx, col_l = data_cols[i]
            prev_l = data_cols[i - 1][1]
            tipo = str(periodos[i].get("tipo_periodo", "")).upper()
            if tipo != "PARCIAL":
                ws.cell(row=4, column=col_idx).value = f"={col_l}6/{prev_l}6-1"
            else:
                # Limpiar celda (el parcial no tiene YoY directo)
                ws.cell(row=4, column=col_idx).value = None

    # Col proyectada: siempre recibe YoY vs col_last
    proj_idx = cols["proj_idx"]
    proj_l = cols["proj"]
    last_l = cols["last"]
    ws.cell(row=4, column=proj_idx).value = f"={proj_l}6/{last_l}6-1"

    # Celda promedio: col_helper (col_proj + 1)
    helper_idx = cols["helper_idx"]
    if len(data_cols) >= 2:
        second_l = data_cols[1][1]
        ws.cell(row=4, column=helper_idx).value = f"=AVERAGE({second_l}4:{proj_l}4)"


# ---------------------------------------------------------------------------
# Inyeccion de datos de periodos individuales
# ---------------------------------------------------------------------------

def inject_estado_resultados(ws, col, periodo):
    er = periodo.get("estado_resultados") or {}
    col_l = get_column_letter(col)

    ingresos = to_float(er.get("ingresos_operativos_netos"))
    costo_ventas = abs(to_float(er.get("costo_de_ventas")))
    gastos_op = abs(to_float(get_alias_value(
        er, "gastos_operativos", "gastos_operativos_totales"
    )))
    gastos_gen = abs(to_float(er.get("gastos_generales")))
    gastos_arr = abs(to_float(er.get("gastos_por_arrendamientos")))
    servicios = abs(to_float(er.get("servicios_externos_y_honorarios")))
    gastos_adm = abs(to_float(get_alias_value(
        er, "gastos_de_administracion", "gastos_de_administration"
    )))
    gastos_vta = abs(to_float(er.get("gastos_de_venta")))
    gastos_per = abs(to_float(er.get("gastos_de_personal")))
    otros_ing_op = abs(to_float(er.get("otros_ingresos_operativos")))
    otros_gtos_op = abs(to_float(er.get("otros_gastos_operativos")))
    otros_gtos_nop = abs(to_float(er.get("otros_gastos_no_operativos")))
    otros_ing_nop = abs(to_float(er.get("otros_ingresos_no_operativos")))
    rif = to_float(er.get("resultado_financiero_neto"))
    isr_dif = abs(to_float(er.get("isr_diferido")))
    isr_cor = abs(to_float(er.get("isr_corriente")))
    ptu = abs(to_float(er.get("provision_ptu")))
    total_imp_gen = abs(to_float(er.get("total_impuestos_generico")))

    desgloses_sum = gastos_adm + gastos_vta + gastos_per + gastos_gen + gastos_arr + servicios
    gastos_op_f13 = max(gastos_op - desgloses_sum, 0.0)
    gastos_gen_f14 = gastos_gen + gastos_arr + servicios
    otros_gtos_ing_f18 = otros_gtos_op - otros_ing_op

    write_cell(ws, 6, col, ingresos)
    write_cell(ws, 8, col, ingresos)
    write_cell(ws, 9, col, costo_ventas)
    set_formula_cell(ws, f"={col_l}8-{col_l}9", row=10, col=col)
    set_formula_cell(ws, f"=SUM({col_l}13:{col_l}18)", row=12, col=col)
    write_cell(ws, 13, col, gastos_op_f13)
    write_cell(ws, 14, col, gastos_gen_f14)
    write_cell(ws, 15, col, gastos_adm)
    ws.cell(row=16, column=3).value = "Gastos de venta"
    ws.cell(row=17, column=3).value = "Gastos de personal"
    ws.cell(row=18, column=3).value = "Otros Gastos/Ingresos Op"
    write_cell(ws, 16, col, gastos_vta)
    write_cell(ws, 17, col, gastos_per)
    write_cell(ws, 18, col, otros_gtos_ing_f18)
    set_formula_cell(ws, f"={col_l}10-{col_l}12", row=19, col=col)
    ws.cell(row=20, column=3).value = "Otros Gastos No Op"
    ws.cell(row=21, column=3).value = "Otros Ingresos No Op"
    write_cell(ws, 20, col, otros_gtos_nop)
    write_cell(ws, 21, col, otros_ing_nop)
    ws.cell(row=22, column=3).value = "EBITDA"
    set_formula_cell(ws, f"={col_l}19+ABS({col_l}98)", row=22, col=col)
    write_cell(ws, 24, col, rif, allow_formula_overwrite=True)
    write_cell(ws, 25, col, isr_dif)
    write_cell(ws, 26, col, isr_cor)
    write_cell(ws, 27, col, ptu)
    set_formula_cell(ws, f"=SUM({col_l}25:{col_l}28)", row=29, col=col)
    if (
        abs(isr_dif) < 1e-9
        and abs(isr_cor) < 1e-9
        and abs(ptu) < 1e-9
        and abs(total_imp_gen) >= 1e-9
    ):
        write_cell(ws, 29, col, total_imp_gen, allow_formula_overwrite=True)
    set_formula_cell(ws, f"={col_l}19+{col_l}24-{col_l}29", row=30, col=col)


def inject_balance_general(ws, col, periodo):
    bg = periodo.get("balance_general") or {}
    activos = bg.get("activos") or {}
    pasivos = bg.get("pasivos") or {}
    capital = bg.get("capital_contable") or {}
    circ = activos.get("circulante") or {}
    no_circ = activos.get("no_circulante") or {}
    pc = pasivos.get("corto_plazo") or {}
    lp = pasivos.get("largo_plazo") or {}

    write_cell(ws, 45, col, to_float(circ.get("efectivo_y_equivalentes")))
    write_cell(ws, 46, col, to_float(circ.get("cuentas_por_cobrar_clientes")))
    write_cell(ws, 47, col, to_float(circ.get("impuestos_a_favor_cp")))
    write_cell(ws, 48, col, to_float(circ.get("otros_activos_circulantes")))
    write_cell(ws, 49, col, to_float(circ.get("deudores_diversos_cp")))
    write_cell(ws, 50, col, to_float(circ.get("pagos_anticipados")))
    ws.cell(row=51, column=3).value = "Inventarios"
    write_cell(ws, 51, col, to_float(circ.get("inventarios")))

    ppe_trans = to_float(no_circ.get("equipo_de_transporte"))
    ppe_comp = to_float(no_circ.get("equipo_de_computo"))
    ppe_mob = to_float(no_circ.get("mobiliario_y_equipo_de_oficina"))
    ppe_dep = abs(to_float(get_alias_value(
        no_circ, "depreciacion_acumulada_historica", "depreciacion_acumulada"
    )))
    ppe_neto = to_float(no_circ.get("propiedad_planta_y_equipo_neto"))

    write_cell(ws, 65, col, ppe_trans)
    write_cell(ws, 66, col, ppe_comp)
    write_cell(ws, 67, col, ppe_mob)
    write_cell(ws, 68, col, -ppe_dep)

    ppe_sum = abs(ppe_trans) + abs(ppe_comp) + abs(ppe_mob) + ppe_dep
    if ppe_sum > 1e-9:
        write_cell(ws, 69, col, 0.0)
    elif ppe_neto > 1e-9:
        ws.cell(row=69, column=3).value = "PPE Neto"
        write_cell(ws, 69, col, ppe_neto)
    else:
        write_cell(ws, 69, col, 0.0)

    ws.cell(row=70, column=3).value = "Activos Intangibles"
    write_cell(ws, 70, col, to_float(no_circ.get("activos_intangibles_neto")))
    ws.cell(row=84, column=3).value = "Activos Diferidos"
    write_cell(ws, 84, col, to_float(no_circ.get("activos_diferidos")))

    pasivo_cp_map = [
        (105, "proveedores", "Proveedores"),
        (106, "impuestos_y_cuotas_por_pagar", "Impuestos y cuotas por pagar"),
        (107, "otros_pasivos_corto_plazo", "Otros Pasivos CP"),
        (108, "acreedores_diversos", "Acreedores diversos"),
        (109, "provisiones", "Provisiones"),
        (110, "anticipo_de_clientes", "Anticipo de Clientes"),
        (111, "deuda_financiera_cp", "Deuda financiera CP"),
    ]
    for row, key, label in pasivo_cp_map:
        ws.cell(row=row, column=3).value = label
        write_cell(ws, row, col, to_float(pc.get(key)))
    for row in range(112, 119):
        write_cell(ws, row, col, 0.0)

    write_cell(ws, 120, col, to_float(lp.get("dividendos_decretados")))
    write_cell(ws, 121, col, to_float(lp.get("pasivo_por_arrendamiento")))
    write_cell(ws, 122, col, to_float(lp.get("deuda_financiera_lp")))

    write_cell(ws, 126, col, to_float(capital.get("capital_social")))
    write_cell(ws, 127, col, to_float(capital.get("utilidades_ejercicios_anteriores")))
    write_cell(ws, 128, col, to_float(capital.get("resultado_del_ejercicio_balance")))


# ---------------------------------------------------------------------------
# Columna proyectada y helper de anualizacion
# ---------------------------------------------------------------------------

def inject_datos_projection(ws, cols, ultimo_periodo):
    """
    Escribe las formulas de la columna proyectada (col_proj) y la
    columna helper (col_helper) en "1. Datos".

    Para periodo PARCIAL:
      col_helper_row = col_last_row / mes_cierre
      col_proj_row   = col_helper_row * 12         (P&L rows)
      Balance rows   : vacios (instantanea, no se proyecta)
    Para periodo ANUAL:
      col_proj_row   = col_last_row               (espejo)
      col_helper     : no se escribe
    Subtotales: se inyectan via inject_native_formulas.
    """
    tipo = str(ultimo_periodo.get("tipo_periodo", "")).upper()
    es_parcial = tipo == "PARCIAL"
    mes = int(ultimo_periodo.get("mes_cierre") or 12)

    last_l = cols["last"]
    proj_idx = cols["proj_idx"]
    proj_l = cols["proj"]
    helper_idx = cols["helper_idx"]
    helper_l = cols["helper"]

    # Subtotales de la col proyectada (mismas formulas que cols historicas)
    inject_native_formulas(ws, proj_idx)

    if es_parcial:
        # Helper: promedio mensual = col_last / mes
        for row in PL_INJECT_ROWS:
            ws.cell(row=row, column=helper_idx).value = f"=+{last_l}{row}/{mes}"
        # Proj: anualizacion = helper * 12
        for row in PL_INJECT_ROWS:
            ws.cell(row=row, column=proj_idx).value = f"=+{helper_l}{row}*12"
        # Fila 98 (depreciacion periodo): tambien anualizamos
        ws.cell(row=98, column=helper_idx).value = f"=+{last_l}98/{mes}"
        ws.cell(row=98, column=proj_idx).value = f"=+{helper_l}98*12"
        # Balance rows en col_proj: vacias
        for row in BAL_INJECT_ROWS:
            ws.cell(row=row, column=proj_idx).value = None
    else:
        # Anual: proyectado = espejo de col_last
        for row in PL_INJECT_ROWS:
            ws.cell(row=row, column=proj_idx).value = f"=+{last_l}{row}"
        ws.cell(row=98, column=proj_idx).value = f"=+{last_l}98"
        # Balance rows en col_proj: espejo
        for row in BAL_INJECT_ROWS:
            ws.cell(row=row, column=proj_idx).value = f"=+{last_l}{row}"
        # Limpiar helper
        for row in PL_INJECT_ROWS + BAL_INJECT_ROWS + [98]:
            ws.cell(row=row, column=helper_idx).value = None


# ---------------------------------------------------------------------------
# CAGR filas 192-213 (Analisis horizontal)
# ---------------------------------------------------------------------------

_CAGR_REF_PAT = re.compile(r"^\s*=\+?([A-Z]+)(\d+)\s*$", re.IGNORECASE)


def _detect_cagr_ref_row(cell_val):
    """Detecta la fila financiera que copia un formula de la forma =+D8 o =D8."""
    if not isinstance(cell_val, str):
        return None
    m = _CAGR_REF_PAT.match(cell_val)
    if m:
        return int(m.group(2))
    return None


def inject_datos_cagr(ws, cols):
    """
    Actualiza las formulas CAGR en las filas de analisis horizontal.

    Para cada fila CAGR:
    - Escribe/extiende formulas de copia de datos =+col_letter{fin_row} para D..col_last.
    - Escribe CAGR en col_proj: =(col_last/col_first)^(1/N)
    - Escribe delta en col_helper: =col_proj-1
    - Limpia celdas stale entre col_last+1 y col_proj-1 (antiguas CAGR/delta del template).
    """
    first_l = cols["first"]
    last_l = cols["last"]
    first_idx = cols["first_idx"]
    last_idx = cols["last_idx"]
    proj_idx = cols["proj_idx"]
    proj_l = cols["proj"]
    helper_idx = cols["helper_idx"]
    helper_l = cols["helper"]
    N = last_idx - first_idx
    if N < 1:
        N = 1

    cagr_rows = list(range(192, 198)) + list(range(205, 214))

    for row in cagr_rows:
        # Detectar fila financiera de referencia desde col D
        d_val = ws.cell(row=row, column=first_idx).value
        fin_row = _detect_cagr_ref_row(d_val)

        if fin_row is None:
            # Si D no tiene formula de copia, intentar detectar desde cualquier col de datos
            for ci, cl in cols["data_cols"]:
                v = ws.cell(row=row, column=ci).value
                fin_row = _detect_cagr_ref_row(v)
                if fin_row is not None:
                    break

        if fin_row is None:
            # No hay referencia detectable; ignorar fila
            continue

        # Extender formulas de copia para TODOS los anios D..col_last
        for col_idx, col_l in cols["data_cols"]:
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = f"=+{col_l}{fin_row}"

        # Limpiar stale entre col_last+1 y col_proj-1
        for stale_col in range(last_idx + 1, proj_idx):
            cell = ws.cell(row=row, column=stale_col)
            if not isinstance(cell, MergedCell):
                cell.value = None

        # Escribir CAGR y delta
        cagr_cell = ws.cell(row=row, column=proj_idx)
        if not isinstance(cagr_cell, MergedCell):
            cagr_cell.value = f"=({last_l}{row}/{first_l}{row})^(1/{N})"
        delta_cell = ws.cell(row=row, column=helper_idx)
        if not isinstance(delta_cell, MergedCell):
            delta_cell.value = f"={proj_l}{row}-1"

    # Encabezados de seccion en col_proj y col_helper
    for hdr_row in [191, 204]:
        c1 = ws.cell(row=hdr_row, column=proj_idx)
        c2 = ws.cell(row=hdr_row, column=helper_idx)
        if not isinstance(c1, MergedCell) and c1.value is None:
            c1.value = "(Crecimiento)"
        if not isinstance(c2, MergedCell) and c2.value is None:
            c2.value = "( % )"


# ---------------------------------------------------------------------------
# Limpieza de columnas obsoletas
# ---------------------------------------------------------------------------

def clear_stale_columns(ws, cols):
    """
    Limpia columnas obsoletas:
    - Para col_helper (row 5/42/102 y fila 4): limpiar labels legacy del template.
    - Para col_helper+1 en adelante: limpiar todos los datos.
    """
    # 1. Limpiar labels/valores legacy en el helper col (no es un col de datos reales)
    helper_idx = cols["helper_idx"]
    for row in HEADER_ROWS:
        ws.cell(row=row, column=helper_idx).value = None
    # fila 4 del helper la maneja inject_datos_yoy_row4; aqui solo limpiamos el row 5 etc.

    # 2. Limpiar todo desde col_helper+1 hasta max_column
    start_clear = helper_idx + 1
    end_clear = ws.max_column
    if start_clear > end_clear:
        return

    cagr_rows = list(range(192, 198)) + list(range(205, 214)) + [191, 204]

    for col in range(start_clear, end_clear + 1):
        for row in HEADER_ROWS:
            ws.cell(row=row, column=col).value = None
        for row in INJECTION_ROWS:
            ws.cell(row=row, column=col).value = None
        ws.cell(row=4, column=col).value = None
        for row in _subtotal_formulas("X", None).keys():
            ws.cell(row=row, column=col).value = None
        for row in cagr_rows:
            ws.cell(row=row, column=col).value = None


def clear_unmapped_historical_cols(ws, cols):
    """
    Limpia columnas D..col_last que NO recibiran datos en esta inyeccion.
    No aplica: en LEFT-ALIGNED siempre usamos D..col_last sin saltos.
    Pero si la nueva inyeccion usa MENOS columnas que la anterior,
    necesitamos limpiar el exceso entre new_last+1 y old_last.
    """
    # El clear_stale_columns ya maneja col_helper+1 en adelante.
    # Aqui tambien limpiamos entre new_last+1 y new_proj-1 (no hay nada ahi).
    pass  # cubierto por clear_stale_columns


# ---------------------------------------------------------------------------
# Backtracking: escaner de referencias en hojas satelite
# ---------------------------------------------------------------------------

_REF_PAT = re.compile(r"'1\. Datos'!([A-Z]+)(\d+)", re.IGNORECASE)


def _scan_sat_refs_for_col(ws_sat, probe_col_letter):
    """
    Escanea ws_sat buscando referencias '1. Datos'!{probe_col_letter}{row}.
    Retorna (row_map, min_sat_col) donde:
      row_map      : list of (sat_row, dados_row)
      min_sat_col  : primer indice de columna en ws_sat donde aparece la ref
    """
    probe = probe_col_letter.upper()
    results = []
    for ws_row in ws_sat.iter_rows():
        for cell in ws_row:
            val = cell.value
            if not isinstance(val, str) or "'1. Datos'" not in val:
                continue
            for m in _REF_PAT.finditer(val):
                if m.group(1).upper() == probe:
                    results.append((cell.row, cell.column, int(m.group(2))))
    if not results:
        return [], None
    min_sat_col = min(r[1] for r in results)
    row_map = [(sat_r, dat_r) for (sat_r, _, dat_r) in results]
    return row_map, min_sat_col


def _patch_sat_all_years(ws_sat, cols, row_map, sat_first_col_idx,
                         last_col_proj_rows=()):
    """
    Escribe formulas en ws_sat para todos los N anios.

    row_map           : list of (sat_row, dados_row)
    sat_first_col_idx : columna en ws_sat que corresponde a dados col_first
    last_col_proj_rows: conjunto de dados_rows donde la ultima columna sat
                        debe apuntar a dados col_proj en vez de col_last
    """
    if not row_map or sat_first_col_idx is None:
        return

    offset = sat_first_col_idx - cols["first_idx"]

    for dados_col_idx, dados_col_l in cols["data_cols"]:
        sat_col_idx = dados_col_idx + offset
        is_last = dados_col_idx == cols["last_idx"]

        for sat_row, dados_row in row_map:
            if is_last and dados_row in last_col_proj_rows:
                ref_col = cols["proj"]
            else:
                ref_col = dados_col_l
            set_formula_cell(
                ws_sat,
                f"='1. Datos'!{ref_col}{dados_row}",
                row=sat_row,
                col=sat_col_idx,
            )


# ---------------------------------------------------------------------------
# repair_calculos_full
# ---------------------------------------------------------------------------

def repair_calculos_full(wb, cols):
    """
    Reconecta '2.Calculos (2)' para todos los N anios.
    Offset: calc_col = dados_col - 1.
    Exception: fila 22 (EBIT) en la ultima calc_col usa dados col_proj.
    """
    calc_name = _find_calculos_sheet_name(wb)
    if not calc_name or "1. Datos" not in wb.sheetnames:
        print("AVISO: no se encontro hoja Calculos o 1. Datos.")
        return

    ws_calc = wb[calc_name]
    ws_datos = wb["1. Datos"]

    # Detectar offset via scan de dados D
    row_map_d, sat_col_d = _scan_sat_refs_for_col(ws_calc, cols["first"])
    if not row_map_d or sat_col_d is None:
        print(f"AVISO: {calc_name} sin refs a 1. Datos!{cols['first']}; usando offset -1.")
        sat_col_d = cols["first_idx"] - 1

    offset = sat_col_d - cols["first_idx"]
    print(f"INFO: {calc_name} offset sat={offset} (sat_col para D={sat_col_d})")

    # El EBIT (dados fila 19) aparece en sat via la referencia dados_row=19
    # En la ultima columna de calc debe usar col_proj en vez de col_last
    ebit_dados_row = 19
    last_col_proj_rows = {ebit_dados_row}

    _patch_sat_all_years(ws_calc, cols, row_map_d, sat_col_d,
                         last_col_proj_rows=last_col_proj_rows)

    # Actualizar encabezado de anios en calc (fila 5)
    _update_sat_year_headers(ws_calc, cols, sat_col_d, header_row=5)

    # Limpiar columnas stale mas alla del rango N
    sat_last_col = cols["last_idx"] + offset
    rows_to_clear = list({sr for sr, _ in row_map_d}) + [5]
    _clear_stale_sat_cols(ws_calc, sat_last_col, rows_to_clear)

    _clean_broken_refs(ws_calc, calc_name)
    print(f"INFO: {calc_name} reparado para {cols['n']} anios.")


# ---------------------------------------------------------------------------
# repair_razones_full
# ---------------------------------------------------------------------------

def repair_razones_full(wb, cols):
    """
    Reconecta 'Razones financieras' para todos los N anios.
    Offset: rf_col = dados_col + 1.
    """
    rf_name = "Razones financieras"
    if rf_name not in wb.sheetnames or "1. Datos" not in wb.sheetnames:
        print(f"AVISO: no se encontro '{rf_name}'.")
        return

    ws_rf = wb[rf_name]
    row_map_d, sat_col_d = _scan_sat_refs_for_col(ws_rf, cols["first"])
    if not row_map_d or sat_col_d is None:
        print(f"AVISO: {rf_name} sin refs a 1. Datos!{cols['first']}; usando offset +1.")
        sat_col_d = cols["first_idx"] + 1

    offset = sat_col_d - cols["first_idx"]
    print(f"INFO: {rf_name} offset sat={offset}")

    _patch_sat_all_years(ws_rf, cols, row_map_d, sat_col_d)

    # Actualizar encabezado de anios en la fila de DATOS
    row_datos = _find_row_by_labels(ws_rf, ["DATOS"], default_row=4, label_col=2)
    _update_sat_year_headers(ws_rf, cols, sat_col_d, header_row=row_datos)

    # Limpiar columnas stale mas alla del rango N
    sat_last_col = cols["last_idx"] + offset
    rows_to_clear = list({sr for sr, _ in row_map_d}) + [row_datos]
    _clear_stale_sat_cols(ws_rf, sat_last_col, rows_to_clear)

    _clean_broken_refs(ws_rf, rf_name)
    print(f"INFO: {rf_name} reparado para {cols['n']} anios.")


# ---------------------------------------------------------------------------
# repair_dupont_full
# ---------------------------------------------------------------------------

def repair_dupont_full(wb, cols):
    """
    Reconecta 'Dupont' para todos los N anios.
    Detecta dinamicamente el offset de la serie (anios 2+) y el col de anio 1.
    """
    dup_name = "Dupont"
    if dup_name not in wb.sheetnames or "1. Datos" not in wb.sheetnames:
        print(f"AVISO: no se encontro '{dup_name}'.")
        return

    ws_dup = wb[dup_name]

    # Obtener datos_col para anio 2 (cols["first_idx"] + 1 = E)
    if cols["n"] > 1:
        col_y2 = get_column_letter(cols["first_idx"] + 1)
        row_map_series, sat_col_series_start = _scan_sat_refs_for_col(ws_dup, col_y2)
    else:
        row_map_series, sat_col_series_start = [], None

    # Detectar sat_col para anio 1 (dados D) via scan directo
    row_map_y1, sat_col_y1 = _scan_sat_refs_for_col(ws_dup, cols["first"])

    if not row_map_series or sat_col_series_start is None:
        # Solo 1 anio o no hay serie detectada: todo offset 0 desde dados D
        if not row_map_y1:
            print(f"AVISO: {dup_name} sin refs detectables; omitiendo.")
            return
        row_map_series = row_map_y1
        sat_col_series_start = sat_col_y1
        series_offset = 0
    else:
        series_offset = sat_col_series_start - (cols["first_idx"] + 1)

    # sat_col_y1: usar el detectado del scan de dados D; fallback: serie - series_offset - 1
    if sat_col_y1 is None:
        sat_col_y1 = sat_col_series_start - series_offset - 1

    # Si year1 no tiene refs propias, usar el row_map de la serie para el row_map
    if not row_map_y1:
        row_map_y1 = row_map_series

    print(f"INFO: {dup_name} series_offset={series_offset}, sat_col_y1={sat_col_y1}")

    # Parchar todos los N anios
    for i, (dados_col_idx, dados_col_l) in enumerate(cols["data_cols"]):
        if i == 0:
            sat_col_idx = sat_col_y1
            rmap = row_map_y1
        else:
            sat_col_idx = sat_col_series_start + (i - 1)
            rmap = row_map_series

        for sat_row, dados_row in rmap:
            set_formula_cell(
                ws_dup,
                f"='1. Datos'!{dados_col_l}{dados_row}",
                row=sat_row,
                col=sat_col_idx,
            )

    # Actualizar encabezado de anios en fila 10 (si la celda acepta formula)
    header_row = 10
    for i, (dados_col_idx, dados_col_l) in enumerate(cols["data_cols"]):
        sat_col_idx = sat_col_y1 if i == 0 else sat_col_series_start + (i - 1)
        cell_hdr = ws_dup.cell(row=header_row, column=sat_col_idx)
        if cell_hdr.value is None or isinstance(cell_hdr.value, (str, int, float)):
            cell_hdr.value = f"='1. Datos'!{dados_col_l}5"

    _clean_broken_refs(ws_dup, dup_name)
    print(f"INFO: {dup_name} reparado para {cols['n']} anios.")


# ---------------------------------------------------------------------------
# repair_resumen_escenario_full
# ---------------------------------------------------------------------------

def repair_resumen_escenario_full(wb, cols):
    """
    Reconecta '4.Resumen-Escenario' para N anios.
    EBIT historico: cols O..T+N-6 (offset +11) referencian dados D..col_last row 19.
    Ultima col escenario (U+offset): dados col_proj row 19.
    AP30: dados col_last row 119.
    AQ60: dados col_(last-1) row 6.
    AT60: dados col_last row 6.
    AN14: dados col_helper (cagr %) row 192.
    """
    esc_name = "4.Resumen-Escenario"
    if esc_name not in wb.sheetnames or "1. Datos" not in wb.sheetnames:
        print(f"AVISO: '{esc_name}' no encontrado.")
        return

    ws_esc = wb[esc_name]
    ws_datos = wb["1. Datos"]

    # Detectar columna de inicio del EBIT historico (fila 11) via scan
    # Buscamos la ref a dados col_first row 19
    row_map_ebit, sat_ebit_start = _scan_sat_refs_for_col(ws_esc, cols["first"])
    if not row_map_ebit or sat_ebit_start is None:
        # Fallback: buscar via row 19 refs manualmente
        for r in ws_esc.iter_rows():
            for cell in r:
                val = cell.value
                if isinstance(val, str) and f"'1. Datos'!{cols['first']}19" in val:
                    sat_ebit_start = cell.column
                    break
            if sat_ebit_start:
                break
        if not sat_ebit_start:
            print(f"AVISO: {esc_name} sin refs detectables; usando offset +11.")
            sat_ebit_start = cols["first_idx"] + 11

    # Filtrar solo la fila de EBIT (row 19 en dados) para la serie historica
    ebit_row_map = [(sr, dr) for sr, dr in row_map_ebit if dr == 19]
    if not ebit_row_map:
        # Puede que la primera ref use row 15 (anomalia del modelo)
        # Forzar a row 19
        ebit_sat_row = 11  # fila 11 en escenario = EBIT
        ebit_row_map = [(ebit_sat_row, 19)]

    # Escribir EBIT historico: dados D..col_last en cols sat
    for i, (dados_col_idx, dados_col_l) in enumerate(cols["data_cols"]):
        sat_col_idx = sat_ebit_start + i
        for sat_row, _ in ebit_row_map:
            set_formula_cell(
                ws_esc,
                f"='1. Datos'!{dados_col_l}19",
                row=sat_row,
                col=sat_col_idx,
            )

    # Ultima col escenario usa col_proj para EBIT
    last_sat_ebit_col = sat_ebit_start + cols["n"]
    for sat_row, _ in ebit_row_map:
        set_formula_cell(
            ws_esc,
            f"='1. Datos'!{cols['proj']}19",
            row=sat_row,
            col=last_sat_ebit_col,
        )

    # Limpiar columnas stale de EBIT mas alla de last_sat_ebit_col
    ebit_sat_rows = [sr for sr, _ in ebit_row_map]
    _clear_stale_sat_cols(ws_esc, last_sat_ebit_col, ebit_sat_rows)

    # AP30: dados col_last row 119 (deuda LP)
    row_deuda = _find_row_by_labels(
        ws_datos, ["Total Pasivo Largo Plazo", "Pasivo LP Total"],
        default_row=119
    )
    set_formula_cell(ws_esc, f"=+'1. Datos'!{cols['last']}{row_deuda}", coord="AP30")

    # AQ60: dados col_(last-1) row 6 (ventas penultimo)
    if cols["n"] >= 2:
        col_prev_l = get_column_letter(cols["last_idx"] - 1)
        set_formula_cell(ws_esc, f"=AP36/'1. Datos'!{col_prev_l}6", coord="AQ60")

    # AT60: dados col_last row 6 (ventas ultimo)
    set_formula_cell(ws_esc, f"='1. Datos'!{cols['last']}6*AR60", coord="AT60")

    # AN14: cagr ventas netas = dados col_helper row 192
    set_formula_cell(ws_esc, f"=+'1. Datos'!{cols['helper']}192", coord="AN14")

    # Nombre empresa
    set_formula_cell(ws_esc, "='1. Datos'!C2", coord="O4")

    _clean_broken_refs(ws_esc, esc_name)
    print(f"INFO: {esc_name} reparado para {cols['n']} anios.")


# ---------------------------------------------------------------------------
# repair_wacc
# ---------------------------------------------------------------------------

def repair_wacc(wb, cols):
    wacc_name = "WACC"
    if wacc_name not in wb.sheetnames or "1. Datos" not in wb.sheetnames:
        print(f"AVISO: '{wacc_name}' no encontrado.")
        return

    ws_wacc = wb[wacc_name]
    ws_datos = wb["1. Datos"]

    last_l = cols["last"]

    # Blindaje de inputs base
    ws_wacc["G4"].value = 0.0417
    ws_wacc["G5"].value = 0.0687
    ws_wacc["C22"].value = 0.0417
    ws_wacc["C23"].value = 1.101955
    ws_wacc["C24"].value = 0.0687
    ws_wacc["C25"].value = 0.0257

    set_formula_cell(ws_wacc, f"=+'1. Datos'!{last_l}119", coord="H30")
    set_formula_cell(ws_wacc, f"=+'1. Datos'!{last_l}125", coord="H31")
    set_formula_cell(ws_wacc, "=SUM(H30+H31)", coord="H32")
    set_formula_cell(ws_wacc, "=+'Estructura de deuda'!J49", coord="I30")
    set_formula_cell(ws_wacc, "=+'Estructura de deuda'!E49", coord="I31")
    set_formula_cell(ws_wacc, "=I30", coord="K24")
    set_formula_cell(ws_wacc, "=I31", coord="K25")
    set_formula_cell(ws_wacc, "=((K25)*(K22)+((K24)*(1-K23)*(K21)))", coord="K26")

    _clean_broken_refs(ws_wacc, wacc_name)
    print(f"INFO: WACC reparado usando col_last={last_l}.")


# ---------------------------------------------------------------------------
# repair_resumen_final
# ---------------------------------------------------------------------------

def repair_resumen_final(wb, cols):
    resumen_name = "RESUMEN"
    if resumen_name not in wb.sheetnames or "1. Datos" not in wb.sheetnames:
        print(f"AVISO: '{resumen_name}' no encontrado.")
        return

    ws_resumen = wb[resumen_name]
    ws_datos = wb["1. Datos"]

    last_l = cols["last"]

    row_capital_total = _find_row_by_labels(
        ws_datos,
        labels=["Capital Total", "Total Capital Contable", "Capital Contable Total"],
        default_row=135,
    )

    ebitda_formula = f"='1. Datos'!{last_l}30"

    # Buscar etiquetas en RESUMEN
    for label_set, default_row in [
        (["EBITDA 2025", "EBITDA ULTIMO AÑO", "EBITDA ULTIMO ANO", "EBITDA"], 6),
    ]:
        row = _find_row_contains_terms(ws_resumen, ["ebitda"], default_row, label_col=1)
        val_col = _find_value_col_for_label_row(ws_resumen, row, label_col=1)
        set_formula_cell(ws_resumen, ebitda_formula, row=row, col=val_col)

    row_vc = _find_row_contains_terms(
        ws_resumen, ["valor neto contable", "capital invertido"], default_row=5, label_col=1
    )

    set_formula_cell(ws_resumen, "='WACC'!K26", coord="B4")
    set_formula_cell(ws_resumen, ebitda_formula, coord="B6")
    set_formula_cell(ws_resumen, ebitda_formula, coord="B12")
    set_formula_cell(
        ws_resumen, f"='1. Datos'!{last_l}{row_capital_total}", coord="B5"
    )
    set_formula_cell(ws_resumen, "='4.Resumen-Escenario'!AP36", coord="B7")

    valor_col = _find_value_col_for_label_row(ws_resumen, row_vc, label_col=1)
    set_formula_cell(
        ws_resumen,
        f"='1. Datos'!{last_l}{row_capital_total}",
        row=row_vc,
        col=valor_col,
    )

    _clean_broken_refs(ws_resumen, resumen_name)
    print(f"INFO: RESUMEN reparado usando col_last={last_l}.")


# ---------------------------------------------------------------------------
# repair_company_name_placeholders
# ---------------------------------------------------------------------------

def repair_company_name_placeholders(wb, company_name):
    empresa = str(company_name or "").strip() or "Sin Nombre"
    replaced = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                text = value.strip()
                if not text or text.startswith("="):
                    continue
                normalized = _normalize_text(text)
                new_value = None
                if "valoracion negocio en" in normalized:
                    new_value = f"Valoracion Negocio En {empresa}"
                elif "ovando" in normalized:
                    if re.search(r"(?i)grupo\s+ovando", text):
                        if re.fullmatch(r"(?is)\s*grupo\s+ovando[^\r\n]*", text):
                            new_value = empresa
                        else:
                            new_value = re.sub(r"(?i)grupo\s+ovando", empresa, text)
                    else:
                        new_value = empresa
                if new_value is not None and new_value != value:
                    cell.value = new_value
                    replaced += 1
    if "1. Datos" in wb.sheetnames:
        wb["1. Datos"]["C2"].value = empresa
    if "RESUMEN" in wb.sheetnames:
        wb["RESUMEN"]["A1"].value = f"Valoracion Negocio En {empresa}"
    print(f"INFO: nombre empresa dinamico aplicado. Celdas actualizadas: {replaced}")


# ---------------------------------------------------------------------------
# Helpers de limpieza y actualizacion de headers satelite
# ---------------------------------------------------------------------------

def _clear_stale_sat_cols(ws_sat, sat_last_col_idx, rows_to_clear):
    """
    Limpia columnas satelite mas alla del ultimo anio inyectado.
    Elimina formulas '1. Datos' stale de inyecciones anteriores con mas anios.
    """
    start = sat_last_col_idx + 1
    end = ws_sat.max_column
    if start > end:
        return
    cleared = 0
    for col in range(start, end + 1):
        for row in rows_to_clear:
            cell = ws_sat.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if isinstance(val, str) and "'1. Datos'" in val:
                cell.value = None
                cleared += 1
    if cleared:
        print(f"INFO: {cleared} celdas stale limpiadas desde col {get_column_letter(start)}.")


def _clean_broken_refs(ws, sheet_name):
    cleaned = 0
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if "#REF!" in cell.value or "[3]" in cell.value or "[1]" in cell.value:
                cell.value = 0.0
                cleaned += 1
    if cleaned:
        print(f"INFO: {cleaned} refs rotas limpiadas en '{sheet_name}'.")


def _update_sat_year_headers(ws_sat, cols, sat_first_col_idx, header_row):
    """Escribe encabezados de anio en la fila header_row de la hoja satelite."""
    offset = sat_first_col_idx - cols["first_idx"]
    for dados_col_idx, dados_col_l in cols["data_cols"]:
        sat_col_idx = dados_col_idx + offset
        ws_sat.cell(row=header_row, column=sat_col_idx).value = (
            f"='1. Datos'!{dados_col_l}5"
        )


# ---------------------------------------------------------------------------
# Funcion principal de inyeccion
# ---------------------------------------------------------------------------

def inyectar_datos_financieros(json_path, template_path, output_path):
    with open(json_path, encoding="utf-8-sig") as f:
        data = json.load(f)

    periodos = data.get("datos_financieros") or []
    empresa = data.get("metadata", {}).get("empresa_detectada", "Sin Nombre")

    if not periodos:
        raise ValueError("El JSON no contiene periodos en 'datos_financieros'.")

    print(f"Empresa : {empresa}")
    print(f"Periodos: {[p.get('anio') for p in periodos]} ({len(periodos)} anios)")

    cols = _build_cols(periodos)
    print(
        f"Columnas: {cols['first']} (anio1) .. {cols['last']} (anioN) | "
        f"Proj={cols['proj']} | Helper={cols['helper']}"
    )

    wb = openpyxl.load_workbook(template_path)
    ws = wb[SHEET_NAME]

    # 1. Limpiar columnas obsoletas de inyecciones anteriores
    clear_stale_columns(ws, cols)

    # 2. Inyectar datos por periodo (LEFT-ALIGNED desde D)
    for i, periodo in enumerate(periodos):
        col_idx, col_l = cols["data_cols"][i]
        anio = periodo.get("anio")
        tipo = str(periodo.get("tipo_periodo", "")).upper()

        # Etiqueta para header rows
        if tipo == "PARCIAL":
            mes = int(periodo.get("mes_cierre") or 12)
            label = f"{anio} ({mes}m)"
        else:
            label = anio

        inject_headers(ws, col_idx, label)
        inject_native_formulas(ws, col_idx)
        inject_estado_resultados(ws, col_idx, periodo)
        inject_balance_general(ws, col_idx, periodo)

    # 3. Headers de anio en filas 5, 42, 102
    inject_datos_year_headers(ws, cols, periodos)

    # 4. Fila 4: crecimiento YoY
    inject_datos_yoy_row4(ws, cols, periodos)

    # 5. Columna proyectada y helper
    inject_datos_projection(ws, cols, periodos[-1])

    # 6. CAGR filas 192-213
    inject_datos_cagr(ws, cols)

    # 7. Reparar hojas satelite
    repair_calculos_full(wb, cols)
    repair_razones_full(wb, cols)
    repair_dupont_full(wb, cols)
    repair_resumen_escenario_full(wb, cols)
    repair_wacc(wb, cols)
    repair_resumen_final(wb, cols)
    repair_company_name_placeholders(wb, empresa)

    wb.save(output_path)

    # Verificacion post-save
    try:
        chk = openpyxl.load_workbook(output_path, data_only=False, read_only=True)
        val_b5 = chk["RESUMEN"]["B5"].value if "RESUMEN" in chk.sheetnames else "N/A"
        print(f"CHECK POST-SAVE RESUMEN!B5 = {val_b5!r}")
        chk.close()
    except Exception as exc:
        print(f"CHECK POST-SAVE ERROR: {exc}")

    print(
        f"Listo. Archivo generado en '{output_path}'. "
        f"{cols['n']} anios inyectados ({cols['first']}..{cols['last']})."
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    import sys
    import os

    json_path = JSON_FILE
    template_path = TEMPLATE_FILE

    args = sys.argv[1:]
    output_path = None
    for arg in args:
        if arg.endswith(".json") and os.path.isfile(arg):
            json_path = arg
        elif arg.endswith(".xlsx") and "template" not in arg.lower():
            output_path = arg

    if output_path is None:
        try:
            with open(json_path, encoding="utf-8-sig") as f:
                meta = json.load(f).get("metadata", {})
            empresa = meta.get("empresa_detectada", "")
        except Exception:
            empresa = ""
        output_path = nombre_a_archivo(empresa or "Valuacion")

    print(f"JSON     : {json_path}")
    print(f"Template : {template_path}")
    print(f"Output   : {output_path}")

    inyectar_datos_financieros(json_path, template_path, output_path)


if __name__ == "__main__":
    main()
